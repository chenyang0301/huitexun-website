#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build iot-cards.json from the three Tencent Docs exports.

Reads the exported .xlsx files (doc1.xlsx / doc2.xlsx / doc3.xlsx), extracts
ICCID + 续费时间 + 到期时间 from the card-level sub-tables, normalizes dates,
de-duplicates by ICCID (keeping the latest expiry, backfilling renewal),
strips all sensitive columns, and writes a compact iot-cards.json.

Usage:
    python build_iot_cards.py
"""
import os
import re
import json
import gzip
import zipfile
import datetime
import hashlib
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SITE = os.path.abspath(os.path.join(HERE, ".."))

# (xlsx file, sheet name, iccid_col, expiry_col, renewal_col, has_header)
SOURCES = [
    ("doc1.xlsx", "总表",             0, 5, 6, True),
    ("doc1.xlsx", "2024.3.8停机列表", 0, 5, 6, True),
    ("doc1.xlsx", "Sheet1",          0, 7, 8, True),
    ("doc1.xlsx", "1年期",           0, 5, 6, True),
    ("doc1.xlsx", "2年",             0, 5, 6, False),
    ("doc1.xlsx", "3年期",           0, 5, 6, True),
    ("doc1.xlsx", "对应信息",         1, 8, None, True),   # 到期用 订单到期时间(8)，卡到期时间(7)常空
    ("doc2.xlsx", "4G联通",          1, 6, 5, True),
    ("doc2.xlsx", "4G移动",          1, None, None, True),
    ("doc2.xlsx", "4G电信",          1, None, None, True),
    ("doc2.xlsx", "5G联通",          1, None, None, True),
    ("doc2.xlsx", "5G移动",          1, None, None, True),
    ("doc2.xlsx", "5G电信",          1, None, None, True),
    ("doc2.xlsx", "好数-测试卡",       1, None, None, True),
    ("doc3.xlsx", "电信卡-3年列表",    1, 6, 5, True),
    ("doc3.xlsx", "电信卡-大流量列表",  1, 6, 5, True),
    ("doc3.xlsx", "联通卡列表",        1, None, None, True),
    ("doc3.xlsx", "移动卡列表",        1, None, None, True),
]

EPOCH = datetime.date(1900, 1, 1)

_EXCEL_EPOCH = datetime.datetime(1899, 12, 30)


def excel_serial_to_date(v):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return ""
    if f < 1 or f > 80000:
        return ""
    return (_EXCEL_EPOCH + datetime.timedelta(days=f)).strftime("%Y-%m-%d")


def norm_date(v):
    """Normalize a cell value to YYYY-MM-DD, or '' if missing/unparseable."""
    if v is None:
        return ""
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "#N/A", "N/A", "-", "/", "无"):
            return ""
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日",
                    "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
                    "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
            try:
                return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
        if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
            return s
        return ""  # unparseable -> drop
    # numeric -> Excel serial date
    return excel_serial_to_date(v)


def parse_dt(s):
    if not s:
        return EPOCH
    try:
        return datetime.datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return EPOCH


def strip_styles(src):
    tmp = src[:-5] + "_ns.xlsx"
    with zipfile.ZipFile(src) as z:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as z2:
            for n in z.namelist():
                if n == "xl/styles.xml":
                    z2.writestr(n, '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                                   '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                                   '</styleSheet>')
                else:
                    z2.writestr(n, z.read(n))
    return tmp


def main():
    records = {}
    stats = {"rows": 0, "kept": 0, "merged": 0, "dup_dropped": 0}
    for fname, sheet, ic, ec, rc, has_header in SOURCES:
        path = os.path.join(HERE, fname)
        if not os.path.exists(path):
            print(f"[skip] {fname} not found")
            continue
        tmp = strip_styles(path)
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            print(f"[skip] sheet '{sheet}' missing in {fname}")
            wb.close()
            continue
        ws = wb[sheet]
        start = 1 if has_header else 0
        n = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < start:
                continue
            if ic >= len(row):
                continue
            raw = row[ic]
            if raw is None:
                continue
            iccid = str(raw).strip().upper()
            if not re.match(r"^[0-9A-Za-z]{10,}$", iccid):
                continue
            # ec / rc may be None in SOURCES, meaning that sheet has no such column
            exp = norm_date(row[ec]) if ec is not None and ec < len(row) else ""
            ren = norm_date(row[rc]) if rc is not None and rc < len(row) else ""
            stats["rows"] += 1
            n += 1
            if iccid not in records:
                records[iccid] = {"iccid": iccid, "renewal_date": ren, "expiry_date": exp}
                stats["kept"] += 1
            else:
                cur = records[iccid]
                if parse_dt(exp) > parse_dt(cur["expiry_date"]):
                    stats["merged"] += 1
                    records[iccid] = {
                        "iccid": iccid,
                        "renewal_date": ren or cur["renewal_date"],
                        "expiry_date": exp,
                    }
                else:
                    stats["dup_dropped"] += 1
                    if not cur["renewal_date"] and ren:
                        cur["renewal_date"] = ren
        wb.close()
        print(f"[ok] {fname}/{sheet}: {n} card rows")

    out = sorted(records.values(), key=lambda r: r["iccid"])
    out_path = os.path.join(SITE, "iot-cards.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))

    # gzip + version file for the query page (7MB plain JSON is slow to load)
    gz_path = os.path.join(SITE, "iot-cards.json.gz")
    with open(out_path, "rb") as f:
        raw = f.read()
    with open(gz_path, "wb") as f:
        f.write(gzip.compress(raw, compresslevel=9))
    version = hashlib.md5(raw).hexdigest()[:12]
    with open(os.path.join(SITE, "iot-cards.version"), "w") as f:
        f.write(version)

    with_exp = sum(1 for r in out if r["expiry_date"])
    with_ren = sum(1 for r in out if r["renewal_date"])
    size = os.path.getsize(out_path)
    gz_size = os.path.getsize(gz_path)
    print(f"\nTotal unique ICCIDs : {len(out)}")
    print(f"  with 到期时间      : {with_exp}")
    print(f"  with 续费时间      : {with_ren}")
    print(f"  rows scanned       : {stats['rows']}")
    print(f"  merged (later exp) : {stats['merged']}")
    print(f"  dup dropped        : {stats['dup_dropped']}")
    print(f"Wrote {out_path} ({size/1024:.0f} KB)")
    print(f"Wrote {gz_path} ({gz_size/1024:.0f} KB, {gz_size*100.0/size:.1f}% of plain)")
    print(f"Wrote iot-cards.version = {version}")


if __name__ == "__main__":
    main()
